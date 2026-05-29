import re
import datetime
from io import BytesIO
from typing import Optional

from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

COLUMN_HINTS: dict[str, list[str]] = {
    "name": [
        "task/công việc", "tính năng chi tiết", "tính năng chính",
        "task", "công việc", "tính năng", "feature", "tên task", "tên", "name",
    ],
    "priority": ["priority", "ưu tiên", "độ ưu tiên"],
    "deadline": ["kết thúc", "end date", "end", "deadline", "ngày kết thúc"],
    "start_date": ["bắt đầu", "start date", "start", "ngày bắt đầu"],
    "assignee": ["owner", "assignee", "người làm", "phụ trách", "người phụ trách"],
    "module": ["module", "phân hệ", "nhóm", "group"],
    "description": ["ghi chú", "notes", "note", "mô tả", "description"],
    "status": ["trạng thái", "status"],
}

PRIORITY_MAP = {
    "critical": "URGENT", "urgent": "URGENT",
    "high": "HIGH", "cao": "HIGH",
    "medium": "MEDIUM", "trung bình": "MEDIUM", "tb": "MEDIUM",
    "low": "LOW", "thấp": "LOW",
}

STATUS_MAP = {
    "in progress": "IN_PROGRESS", "đang làm": "IN_PROGRESS", "inprogress": "IN_PROGRESS",
    "to do": "TODO", "todo": "TODO", "plan": "TODO", "planned": "TODO",
    "chưa làm": "TODO", "chưa bắt đầu": "TODO",
    "done": "DONE", "xong": "DONE", "hoàn thành": "DONE",
    "review": "DONE", "đang review": "DONE",
    "cancelled": "CANCELLED", "canceled": "CANCELLED", "hủy": "CANCELLED", "đã hủy": "CANCELLED",
}

_EMOJI_RE = re.compile(
    "[\U0001F300-\U0001F9FF\U00002600-\U000027BF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF]",
    re.UNICODE,
)

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB


class PreviewRow(BaseModel):
    row_index: int
    name: str
    priority: str = "MEDIUM"
    deadline: Optional[str] = None
    start_date: Optional[str] = None
    assignee_name: Optional[str] = None
    module: Optional[str] = None
    description: Optional[str] = None
    status: str = "TODO"


class ParseResult(BaseModel):
    sheets: list[str]
    active_sheet: str
    column_map: dict
    rows: list[PreviewRow]


class ImportConfirmRow(BaseModel):
    row_index: int
    name: str
    priority: str = "MEDIUM"
    deadline: Optional[str] = None
    start_date: Optional[str] = None
    assignee_name: Optional[str] = None
    module: Optional[str] = None
    description: Optional[str] = None
    status: str = "TODO"
    skip: bool = False


class ImportConfirmBody(BaseModel):
    rows: list[ImportConfirmRow]


def normalize_priority(raw: str) -> str:
    if not raw:
        return "MEDIUM"
    cleaned = _EMOJI_RE.sub("", str(raw)).strip().lower()
    for key, val in PRIORITY_MAP.items():
        if key in cleaned:
            return val
    return "MEDIUM"


def normalize_status(raw: str) -> str:
    if not raw:
        return "TODO"
    cleaned = str(raw).strip().lower()
    for key, val in STATUS_MAP.items():
        if key in cleaned:
            return val
    return "TODO"


def detect_column_map(headers: list[str]) -> dict[str, Optional[int]]:
    result: dict[str, Optional[int]] = {k: None for k in COLUMN_HINTS}
    for field, hints in COLUMN_HINTS.items():
        for i, h in enumerate(headers):
            h_lower = str(h).lower().strip()
            if any(hint in h_lower for hint in hints):
                result[field] = i
                break
    return result


def _cell_str(cell) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def _cell_date(cell, default_year: Optional[int] = None) -> Optional[str]:
    if cell is None or cell.value is None:
        return None
    v = cell.value
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    if default_year:
        for fmt in ("%d/%m", "%d-%m"):
            try:
                parsed = datetime.datetime.strptime(s, fmt)
                return datetime.date(default_year, parsed.month, parsed.day).isoformat()
            except ValueError:
                pass
    return None


def parse_xlsx(file_bytes: bytes, sheet_name: Optional[str] = None, filename: str = "") -> ParseResult:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed — add it to requirements.txt")
    if len(file_bytes) > MAX_FILE_SIZE:
        raise ValueError("File quá lớn (tối đa 5 MB)")

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    default_year = _infer_year(filename, wb)

    sheet_names = [s for s in wb.sheetnames if wb[s].sheet_state != "hidden"]
    if not sheet_names:
        sheet_names = list(wb.sheetnames)

    if sheet_name and sheet_name in wb.sheetnames:
        return _parse_sheet(wb[sheet_name], sheet_names, sheet_name, default_year)

    candidates = [
        _parse_sheet(wb[name], sheet_names, name, default_year)
        for name in sheet_names
    ]
    for result in candidates:
        if result.rows:
            return result
    return candidates[0] if candidates else ParseResult(sheets=[], active_sheet="", column_map={}, rows=[])


def _parse_sheet(ws, sheet_names: list[str], active: str, default_year: Optional[int]) -> ParseResult:
    # Find first row that looks like a header (≥ 2 non-empty string cells)
    header_row_idx: Optional[int] = None
    headers: list[str] = []
    for row_num, row in enumerate(ws.iter_rows(), start=1):
        str_cells = [c for c in row if c.value and isinstance(c.value, str) and c.value.strip()]
        if len(str_cells) >= 2:
            header_row_idx = row_num
            headers = [_cell_str(c) for c in row]
            break

    if header_row_idx is None:
        return ParseResult(sheets=sheet_names, active_sheet=active, column_map={}, rows=[])

    col_map = detect_column_map(headers)

    rows: list[PreviewRow] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1), start=header_row_idx + 1):
        row_list = list(row)

        name_idx = col_map.get("name")
        if name_idx is None or name_idx >= len(row_list):
            continue
        name_val = _cell_str(row_list[name_idx])
        if not name_val or name_val.lower() in ("nan", "none", "n/a", ""):
            continue

        def _get(field: str) -> Optional[str]:
            idx = col_map.get(field)
            if idx is None or idx >= len(row_list):
                return None
            v = _cell_str(row_list[idx])
            return v if v else None

        def _get_date(field: str) -> Optional[str]:
            idx = col_map.get(field)
            if idx is None or idx >= len(row_list):
                return None
            return _cell_date(row_list[idx], default_year=default_year)

        rows.append(PreviewRow(
            row_index=row_num,
            name=name_val,
            priority=normalize_priority(_get("priority") or ""),
            deadline=_get_date("deadline"),
            start_date=_get_date("start_date"),
            assignee_name=_get("assignee"),
            module=_get("module"),
            description=_get("description"),
            status=normalize_status(_get("status") or ""),
        ))

    return ParseResult(
        sheets=sheet_names,
        active_sheet=active,
        column_map={k: v for k, v in col_map.items() if v is not None},
        rows=rows,
    )


def _infer_year(filename: str, wb) -> Optional[int]:
    text_parts = [filename]
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 20, 20), values_only=True):
            text_parts.extend(str(value) for value in row if value is not None)
    text = " ".join(text_parts)
    matches = [int(match) for match in re.findall(r"\b(20\d{2})\b", text)]
    return matches[0] if matches else None


async def _resolve_assignee(name: str, db: AsyncSession) -> Optional[int]:
    row = (await db.execute(
        text("SELECT id FROM users WHERE LOWER(full_name) ILIKE LOWER(:name) LIMIT 1"),
        {"name": f"%{name}%"},
    )).fetchone()
    return row[0] if row else None


async def bulk_create_tasks(
    project_id: int,
    rows: list[ImportConfirmRow],
    user_id: int,
    db: AsyncSession,
) -> dict:
    created = 0
    errors: list[dict] = []
    project_row = (await db.execute(
        text("SELECT company_id FROM projects WHERE id = :pid"),
        {"pid": project_id},
    )).fetchone()
    company_id = project_row[0] if project_row else None

    for row in rows:
        if row.skip:
            continue
        try:
            assignee_id: Optional[int] = None
            if row.assignee_name:
                assignee_id = await _resolve_assignee(row.assignee_name, db)

            # asyncpg requires datetime.date objects, never strings
            deadline = datetime.date.fromisoformat(row.deadline) if row.deadline else None
            end_at = datetime.date.fromisoformat(row.start_date) if row.start_date else None

            desc = row.description or ""
            if row.module:
                desc = f"[{row.module}] {desc}".strip()

            await db.execute(
                text("""
                    INSERT INTO tasks (
                        name, status, priority, deadline, end_at, description,
                        project_id, assignee_id, company_id, updated_at
                    ) VALUES (
                        :name,
                        CAST(:status AS "TaskStatus"),
                        CAST(:priority AS "Priority"),
                        :deadline, :end_at, :description,
                        :project_id, :assignee_id, :company_id, NOW()
                    )
                """),
                {
                    "name": row.name,
                    "status": row.status,
                    "priority": row.priority,
                    "deadline": deadline,
                    "end_at": end_at,
                    "description": desc or None,
                    "project_id": project_id,
                    "assignee_id": assignee_id,
                    "company_id": company_id,
                },
            )
            created += 1
        except Exception as e:
            errors.append({"row_index": row.row_index, "message": str(e)})

    if created > 0:
        # Update task_count on project
        await db.execute(
            text("UPDATE projects SET task_count = task_count + :n, updated_at = NOW() WHERE id = :pid"),
            {"n": created, "pid": project_id},
        )
        await db.commit()

    return {"created": created, "errors": errors}
