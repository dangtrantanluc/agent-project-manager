import asyncio
import datetime
import json
import logging
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from ai_agent.planning.planning_agent import PlanningAgent
from app.modules.tasks.import_service import MAX_FILE_SIZE, normalize_priority, normalize_status

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

VALID_TASK_STATUSES = {"TODO", "IN_PROGRESS", "DONE", "CANCELLED"}
VALID_PRIORITIES = {"LOW", "MEDIUM", "HIGH", "URGENT"}


class AiImportMilestone(BaseModel):
    tempId: str
    name: str
    description: Optional[str] = None
    dueDate: Optional[str] = None
    status: Optional[str] = None
    skip: bool = False


class AiImportTask(BaseModel):
    tempMilestoneId: Optional[str] = None
    name: str
    description: Optional[str] = None
    priority: str = "MEDIUM"
    status: str = "TODO"
    deadline: Optional[str] = None
    assigneeName: Optional[str] = None
    sourceSheet: Optional[str] = None
    skip: bool = False


class AiImportPreview(BaseModel):
    projectName: str = ""
    summary: str = ""
    milestones: list[AiImportMilestone] = Field(default_factory=list)
    tasks: list[AiImportTask] = Field(default_factory=list)


class AiImportConfirmBody(BaseModel):
    projectName: Optional[str] = None
    summary: Optional[str] = None
    milestones: list[AiImportMilestone] = Field(default_factory=list)
    tasks: list[AiImportTask] = Field(default_factory=list)


def parse_ai_import_xlsx(file_bytes: bytes, filename: str = "") -> AiImportPreview:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed — add it to requirements.txt")
    if len(file_bytes) > MAX_FILE_SIZE:
        raise ValueError("File quá lớn (tối đa 5 MB)")

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    year = _infer_year(filename, wb)
    preview = _deterministic_preview(wb, year)

    if not preview.tasks:
        raise ValueError("Không tìm thấy task phù hợp trong file Excel")
    return preview


async def enrich_with_planning_agent(preview: AiImportPreview, workbook_context: str) -> AiImportPreview:
    """
    Best-effort AI enrichment. If the configured LLM is unavailable, keep the
    deterministic preview so import remains usable in dev/test.
    """
    try:
        agent = PlanningAgent()
        prompt = (
            "Hãy tinh chỉnh kế hoạch import sau. Giữ nguyên JSON, chỉ cải thiện tên milestone, "
            "description ngắn và gom task đúng milestone. Không bịa thêm task mới.\n\n"
            f"Workbook context:\n{workbook_context[:12000]}\n\n"
            f"Current preview JSON:\n{preview.model_dump_json()}"
        )
        plan = await asyncio.wait_for(agent.generate_project_plan(prompt), timeout=12)
        if plan.milestones:
            return _preview_from_project_plan(plan, preview)
    except Exception as exc:
        logger.warning("AI planning import enrichment skipped: %s", exc)
    return preview


def serialize_workbook_context(file_bytes: bytes, max_rows_per_sheet: int = 40) -> str:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    chunks: list[str] = []
    for ws in wb.worksheets:
        if ws.sheet_state == "hidden":
            continue
        chunks.append(f"# Sheet: {ws.title}")
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [_clean_cell(v) for v in row]
            values = [v for v in values if v]
            if values:
                chunks.append(f"{row_index}: " + " | ".join(values[:10]))
            if row_index >= max_rows_per_sheet:
                break
    return "\n".join(chunks)


async def confirm_ai_import(
    project_id: int,
    body: AiImportConfirmBody,
    db: AsyncSession,
) -> dict:
    created_milestones = 0
    created_tasks = 0
    errors: list[dict] = []
    milestone_id_by_temp: dict[str, int] = {}
    skipped_milestone_ids = {m.tempId for m in body.milestones if m.skip}
    project_row = (await db.execute(
        text("SELECT company_id FROM projects WHERE id = :pid"),
        {"pid": project_id},
    )).fetchone()
    company_id = project_row[0] if project_row else None

    for milestone in body.milestones:
        if milestone.skip:
            continue
        try:
            row = (await db.execute(
                text("""
                    INSERT INTO milestones (name, status, due_date, description, project_id, updated_at)
                    VALUES (:name, :status, :due_date, :description, :project_id, NOW())
                    RETURNING id
                """),
                {
                    "name": milestone.name,
                    "status": milestone.status,
                    "due_date": _date_or_none(milestone.dueDate),
                    "description": milestone.description,
                    "project_id": project_id,
                },
            )).fetchone()
            milestone_id_by_temp[milestone.tempId] = row[0]
            created_milestones += 1
        except Exception as exc:
            errors.append({"type": "milestone", "name": milestone.name, "message": str(exc)})

    for task in body.tasks:
        if task.skip or (task.tempMilestoneId and task.tempMilestoneId in skipped_milestone_ids):
            continue
        try:
            assignee_id = await _resolve_assignee(task.assigneeName, db) if task.assigneeName else None
            milestone_id = milestone_id_by_temp.get(task.tempMilestoneId or "")
            await db.execute(
                text("""
                    INSERT INTO tasks (
                        name, status, priority, deadline, description,
                        project_id, assignee_id, milestone_id, company_id, updated_at
                    ) VALUES (
                        :name, CAST(:status AS "TaskStatus"), CAST(:priority AS "Priority"),
                        :deadline, :description, :project_id, :assignee_id, :milestone_id, :company_id, NOW()
                    )
                """),
                {
                    "name": task.name,
                    "status": _normalize_task_status(task.status),
                    "priority": _normalize_task_priority(task.priority),
                    "deadline": _date_or_none(task.deadline),
                    "description": _task_description(task),
                    "project_id": project_id,
                    "assignee_id": assignee_id,
                    "milestone_id": milestone_id,
                    "company_id": company_id,
                },
            )
            created_tasks += 1
        except Exception as exc:
            errors.append({"type": "task", "name": task.name, "message": str(exc)})

    await _recompute_project_counts(project_id, db)
    await _recompute_milestones(project_id, db)
    await db.commit()
    return {"createdMilestones": created_milestones, "createdTasks": created_tasks, "errors": errors}


def _deterministic_preview(wb: Any, year: int) -> AiImportPreview:
    milestones: list[AiImportMilestone] = []
    tasks: list[AiImportTask] = []
    milestone_by_key: dict[str, str] = {}
    project_name = ""
    summary = ""

    def ensure_milestone(name: str, description: str = "", due_date: Optional[str] = None, status: Optional[str] = None) -> str:
        key = _slug(name)
        if key in milestone_by_key:
            return milestone_by_key[key]
        temp_id = f"m{len(milestones) + 1}"
        milestone_by_key[key] = temp_id
        milestones.append(AiImportMilestone(
            tempId=temp_id,
            name=name[:160],
            description=description[:500] or None,
            dueDate=due_date,
            status=status,
        ))
        return temp_id

    for ws in wb.worksheets:
        if ws.sheet_state == "hidden":
            continue
        rows = list(ws.iter_rows(values_only=True))
        non_empty = [[_clean_cell(v) for v in row] for row in rows]
        non_empty = [[v for v in row if v] for row in non_empty if any(row)]
        if not non_empty:
            continue

        if not project_name:
            project_name = _extract_project_name(non_empty)
        if not summary and non_empty:
            summary = " | ".join(non_empty[0][:3])

        if _looks_like_overview(ws.title):
            _parse_overview_rows(non_empty, year, ensure_milestone)
            continue

        milestone_name = _sheet_milestone_name(ws.title, non_empty)
        milestone_desc = " | ".join(non_empty[0][:2]) if non_empty else ""
        sheet_milestone_id = ensure_milestone(milestone_name, milestone_desc, _sheet_due_date(non_empty, year), None)
        header_index, header = _find_header(non_empty)
        if header_index is None:
            continue

        col = _column_indexes(header)
        current_module = ""
        for row in non_empty[header_index + 1:]:
            if _is_group_row(row):
                current_module = row[0]
                continue
            name = _pick(row, col, "name") or _pick(row, col, "feature") or _pick(row, col, "detail")
            detail = _pick(row, col, "detail")
            if not name or len(name) < 3:
                continue
            task_name = name if name != detail else name[:180]
            desc_parts = [p for p in [detail if detail != task_name else "", _pick(row, col, "technical"), current_module] if p]
            tasks.append(AiImportTask(
                tempMilestoneId=sheet_milestone_id,
                name=task_name[:180],
                description=" | ".join(desc_parts)[:1000] or None,
                priority=_normalize_task_priority(_pick(row, col, "priority")),
                status=_normalize_task_status(_pick(row, col, "status")),
                deadline=_parse_date(_pick(row, col, "deadline"), year),
                assigneeName=_pick(row, col, "assignee"),
                sourceSheet=ws.title,
            ))

    return AiImportPreview(
        projectName=project_name or "Imported Project Plan",
        summary=summary[:1000],
        milestones=milestones,
        tasks=tasks[:300],
    )


def _parse_overview_rows(rows: list[list[str]], year: int, ensure_milestone) -> None:
    header_index, header = _find_header(rows)
    if header_index is None:
        return
    col = _column_indexes(header)
    for row in rows[header_index + 1:]:
        sprint = _pick(row, col, "sprint") or _pick(row, col, "name")
        if not sprint:
            continue
        if not re.search(r"^\s*(sprint|giai đoạn|giai doan|phase)\b", sprint, re.IGNORECASE):
            continue
        focus = _pick(row, col, "focus")
        deliverable = _pick(row, col, "deliverable")
        due = _last_date_in_text(" ".join(row), year)
        status = _milestone_status(_pick(row, col, "status"))
        ensure_milestone(sprint, " | ".join(p for p in [focus, deliverable] if p), due, status)


def _preview_from_project_plan(plan: Any, fallback: AiImportPreview) -> AiImportPreview:
    milestones: list[AiImportMilestone] = []
    tasks: list[AiImportTask] = []
    for m_idx, milestone in enumerate(plan.milestones, start=1):
        temp_id = f"m{m_idx}"
        milestones.append(AiImportMilestone(
            tempId=temp_id,
            name=milestone.name,
            description=milestone.goal,
            status="TODO",
        ))
        for task in milestone.tasks:
            tasks.append(AiImportTask(
                tempMilestoneId=temp_id,
                name=task.title,
                description=task.description,
                priority=_normalize_task_priority(task.priority),
                status="TODO",
                assigneeName=task.role,
            ))
    if not tasks:
        return fallback
    if len(tasks) < max(3, int(len(fallback.tasks) * 0.7)):
        logger.warning(
            "AI planning import enrichment ignored because it reduced tasks from %s to %s",
            len(fallback.tasks),
            len(tasks),
        )
        return fallback
    return AiImportPreview(
        projectName=plan.project_name or fallback.projectName,
        summary=plan.summary or fallback.summary,
        milestones=milestones,
        tasks=tasks,
    )


async def _resolve_assignee(name: str, db: AsyncSession) -> Optional[int]:
    row = (await db.execute(
        text("SELECT id FROM users WHERE LOWER(full_name) ILIKE LOWER(:name) LIMIT 1"),
        {"name": f"%{name}%"},
    )).fetchone()
    return row[0] if row else None


async def _recompute_project_counts(project_id: int, db: AsyncSession) -> None:
    await db.execute(
        text("""
            UPDATE projects p SET
                task_count = (SELECT COUNT(*) FROM tasks WHERE project_id = :pid),
                milestone_count = (SELECT COUNT(*) FROM milestones WHERE project_id = :pid),
                updated_at = NOW()
            WHERE p.id = :pid
        """),
        {"pid": project_id},
    )


async def _recompute_milestones(project_id: int, db: AsyncSession) -> None:
    await db.execute(
        text("""
            UPDATE milestones m SET
                task_count = stats.task_count,
                done_count = stats.done_count,
                completion_pct = CASE WHEN stats.task_count = 0 THEN 0
                    ELSE ROUND(stats.done_count * 100.0 / stats.task_count)::int END,
                updated_at = NOW()
            FROM (
                SELECT m.id, COUNT(t.id)::int AS task_count,
                       COUNT(t.id) FILTER (WHERE t.status = 'DONE'::"TaskStatus")::int AS done_count
                FROM milestones m
                LEFT JOIN tasks t ON t.milestone_id = m.id
                WHERE m.project_id = :pid
                GROUP BY m.id
            ) stats
            WHERE m.id = stats.id
        """),
        {"pid": project_id},
    )


def _column_indexes(headers: list[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, header in enumerate(headers):
        h = header.lower()
        if "sprint" in h or "giai đoạn" in h:
            result["sprint"] = idx
        if "trọng tâm" in h or "focus" in h:
            result["focus"] = idx
        if "deliverable" in h or "mốc" in h:
            result["deliverable"] = idx
        if "task" in h or "công việc" in h:
            result["name"] = idx
        if "tính năng chính" in h:
            result["feature"] = idx
        if "tính năng chi tiết" in h:
            result["detail"] = idx
        if "technical" in h or "module /" in h:
            result["technical"] = idx
        if "owner" in h or "phụ trách" in h:
            result["assignee"] = idx
        if "kết thúc" in h or "deadline" in h:
            result["deadline"] = idx
        if "priority" in h or "ưu tiên" in h:
            result["priority"] = idx
        if "trạng thái" in h or "status" in h:
            result["status"] = idx
    return result


def _find_header(rows: list[list[str]]) -> tuple[Optional[int], list[str]]:
    for idx, row in enumerate(rows[:30]):
        joined = " ".join(row).lower()
        if (
            ("task" in joined or "công việc" in joined or "tính năng" in joined or "sprint" in joined)
            and len(row) >= 3
        ):
            return idx, row
    return None, []


def _pick(row: list[str], col: dict[str, int], field: str) -> str:
    idx = col.get(field)
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.date().isoformat() if isinstance(value, datetime.datetime) else value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def _infer_year(filename: str, wb: Any) -> int:
    text = filename + " " + " ".join(wb.sheetnames)
    for ws in wb.worksheets[:2]:
        for row in ws.iter_rows(values_only=True):
            text += " " + " ".join(_clean_cell(v) for v in row if v)
            if len(text) > 3000:
                break
    match = re.search(r"20\d{2}", text)
    return int(match.group(0)) if match else datetime.date.today().year


def _extract_project_name(rows: list[list[str]]) -> str:
    for row in rows[:12]:
        joined = " ".join(row)
        if "Dự án" in joined and len(row) >= 2:
            return row[1]
        if "CRM" in joined or "ODOO" in joined or "LOGISTICS" in joined:
            return joined[:180]
    return ""


def _looks_like_overview(sheet_name: str) -> bool:
    lowered = sheet_name.lower()
    return "overview" in lowered or "tổng quan" in lowered or "tong quan" in lowered


def _sheet_milestone_name(sheet_name: str, rows: list[list[str]]) -> str:
    sprint_match = re.search(r"\b(SPRINT|GIAI ĐOẠN|GIAI DOAN|PHASE)\s*\d+", sheet_name, re.IGNORECASE)
    if sprint_match:
        return sprint_match.group(0).upper()
    if rows and rows[0]:
        first = rows[0][0]
        sprint_match = re.search(r"\b(SPRINT|GIAI ĐOẠN|GIAI DOAN|PHASE)\s*\d+", first, re.IGNORECASE)
        if sprint_match:
            return sprint_match.group(0).upper()
        if "SPRINT" in first.upper() or "GIAI ĐOẠN" in first.upper():
            return first.split("|")[0].strip()
    return re.sub(r"^[^\wÀ-ỹ]+", "", sheet_name).strip()


def _sheet_due_date(rows: list[list[str]], year: int) -> Optional[str]:
    for row in rows[:4]:
        due = _last_date_in_text(" ".join(row), year)
        if due:
            return due
    return None


def _is_group_row(row: list[str]) -> bool:
    if len(row) == 1:
        value = row[0].upper()
        return value.startswith("MODULE") or re.match(r"^[A-Z]\.", value) is not None
    return False


def _parse_date(raw: str, year: int) -> Optional[str]:
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass
    match = re.search(r"(\d{1,2})[/-](\d{1,2})(?:[/-](20\d{2}))?", raw)
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        parsed_year = int(match.group(3) or year)
        try:
            return datetime.date(parsed_year, month, day).isoformat()
        except ValueError:
            return None
    return None


def _last_date_in_text(raw: str, year: int) -> Optional[str]:
    matches = list(re.finditer(r"(\d{1,2})[/-](\d{1,2})(?:[/-](20\d{2}))?", raw))
    if not matches:
        return None
    return _parse_date(matches[-1].group(0), year)


def _date_or_none(raw: Optional[str]) -> Optional[datetime.date]:
    if not raw:
        return None
    return datetime.date.fromisoformat(raw)


def _normalize_task_status(raw: Optional[str]) -> str:
    normalized = normalize_status(raw or "")
    if normalized not in VALID_TASK_STATUSES:
        if "review" in (raw or "").lower():
            return "DONE"
        return "TODO"
    return normalized


def _normalize_task_priority(raw: Optional[str]) -> str:
    normalized = normalize_priority(raw or "")
    return normalized if normalized in VALID_PRIORITIES else "MEDIUM"


def _milestone_status(raw: str) -> Optional[str]:
    status = _normalize_task_status(raw)
    if status == "DONE":
        return "DONE"
    if status == "IN_PROGRESS":
        return "IN_PROGRESS"
    if status == "TODO":
        return "PLANNED"
    return None


def _task_description(task: AiImportTask) -> Optional[str]:
    parts = []
    if task.description:
        parts.append(task.description)
    if task.sourceSheet:
        parts.append(f"Source: {task.sourceSheet}")
    return "\n".join(parts) if parts else None


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
